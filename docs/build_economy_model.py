"""Builds docs/economy_model.xlsx — formula-driven from the Peg + Earn tabs.
Also emits a Values tab (hardcoded results) and economy_model.csv, because openpyxl
writes formulas without cached values and Apple Numbers renders those cells blank."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv, os

HERE = os.path.dirname(os.path.abspath(__file__))
H   = Font(bold=True, color="FFFFFF", size=11)
HF  = PatternFill("solid", fgColor="1F4E3D")          # felt green
IN  = PatternFill("solid", fgColor="FFF3CD")          # yellow = input you own
OUT = PatternFill("solid", fgColor="EAF3EF")          # green  = computed
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="D0D7D3")]*4)

wb = Workbook()

def sheet(name, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    return ws

def head(ws, row, cells, widths=None):
    for i, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font, c.fill, c.alignment = H, HF, Alignment(horizontal="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

# ── PEG ────────────────────────────────────────────────────────────────────────
ws = sheet("Peg", first=True)
ws["A1"] = "THE PEG — change B2 and every tab recomputes"; ws["A1"].font = Font(bold=True, size=13)
ws["A2"], ws["B2"] = "Coins per ₹1 of discount value", 10
ws["B2"].fill, ws["B2"].font = IN, BOLD
ws["A4"] = "Yellow cells are inputs I own and must defend. Green cells are computed."
ws["A5"] = "Referenced everywhere as Peg!$B$2 — never restate the peg in another cell."
ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 12

# ── EARN ───────────────────────────────────────────────────────────────────────
ws = sheet("Earn")
ws["A1"] = "EARN RATES"; ws["A1"].font = Font(bold=True, size=13)
head(ws, 3, ["Event", "Coins"], [34, 12])
for r, (k, v) in enumerate([("Match win", 50), ("Match loss", 15), ("Daily streak", 40),
                            ("Assumed win rate", 0.5)], start=4):
    ws.cell(row=r, column=1, value=k)
    c = ws.cell(row=r, column=2, value=v); c.fill = IN
ws["B7"].number_format = "0%"

head(ws, 9, ["Archetype", "Matches/day", "Days played /wk", "Share of DAU", "Coins/day",
             "Coins/week", "48h budget", "3-week budget"], [16, 13, 15, 13, 12, 12, 12, 14])
# Matches/day is already averaged across all 7 days. The streak bonus is only earned on days
# actually played, so it is weighted by daysPlayed/7 — a Dipper does not get a daily streak daily.
for r, (name, mpd, dpw, share) in enumerate([("Grinder", 15, 7, .25), ("Dipper", 1.43, 2.5, .60),
                                             ("Whale", 6, 7, .15)], start=10):
    ws.cell(row=r, column=1, value=name).font = BOLD
    ws.cell(row=r, column=2, value=mpd).fill = IN
    ws.cell(row=r, column=3, value=dpw).fill = IN
    c = ws.cell(row=r, column=4, value=share); c.fill, c.number_format = IN, "0%"
    for col, f in [(5, f"=B{r}*($B$4*$B$7+$B$5*(1-$B$7))+$B$6*(C{r}/7)"),
                   (6, f"=E{r}*7"), (7, f"=E{r}*2"), (8, f"=E{r}*21")]:
        c = ws.cell(row=r, column=col, value=f); c.fill, c.number_format = OUT, "#,##0"
ws["A14"] = "Blended platform coins/day"
ws["B14"] = "=SUMPRODUCT(E10:E12,D10:D12)"; ws["B14"].fill, ws["B14"].number_format = OUT, "#,##0"

# ── CATALOG ────────────────────────────────────────────────────────────────────
CAT = [
    (1,  "Zepto ₹10 off ₹99",        "voucher",          10,   1.0),
    (2,  "Swiggy ₹40 off ₹149",      "voucher",          40,   1.0),
    (3,  "Zomato ₹50 off ₹199",      "voucher",          50,   1.0),
    (4,  "Starbucks ₹100 off ₹299",  "voucher",         100,   1.0),
    (5,  "PharmEasy ₹200 off ₹799",  "voucher",         200,   1.0),
    (6,  "Decathlon ₹300 off ₹999",  "voucher",         300,   1.0),
    (7,  "Spotify Premium 1mo",      "product·digital", 139,   0.40),
    (8,  "Netflix Mobile 1mo",       "product·digital", 149,   0.40),
    (9,  "JioHotstar Super 1yr",     "product·digital",1499,   0.40),
    (10, "boAt Airdopes 141",        "product·physical",1299,  0.40),
    (11, "Puma Softride sliders",    "product·physical",1499,  0.40),
    (12, "Sony WH-CH520",            "product·physical",2999,  0.40),
]
ws = sheet("Catalog")
ws["A1"] = "CATALOG — coins = MRP × cap × peg.  Cash = MRP − coins/peg."
ws["A1"].font = Font(bold=True, size=13)
head(ws, 3, ["#", "SKU", "Type", "MRP ₹", "Coin cap", "Coins", "Cash ₹",
             "Days: Grinder", "Days: Dipper", "Days: Whale"],
     [4, 30, 17, 10, 10, 10, 10, 13, 13, 13])
for i, (n, name, typ, mrp, cap) in enumerate(CAT):
    r = 4 + i
    ws.cell(row=r, column=1, value=n)
    ws.cell(row=r, column=2, value=name)
    ws.cell(row=r, column=3, value=typ)
    ws.cell(row=r, column=4, value=mrp).fill = IN
    c = ws.cell(row=r, column=5, value=cap); c.fill, c.number_format = IN, "0%"
    c = ws.cell(row=r, column=6, value=f"=D{r}*E{r}*Peg!$B$2"); c.fill, c.number_format = OUT, "#,##0"
    c = ws.cell(row=r, column=7, value=f"=D{r}-F{r}/Peg!$B$2");  c.fill, c.number_format = OUT, "#,##0"
    for col, earn in [(8, "Earn!$E$10"), (9, "Earn!$E$11"), (10, "Earn!$E$12")]:
        c = ws.cell(row=r, column=col, value=f"=F{r}/{earn}")
        c.fill, c.number_format = OUT, "0.0"
    for col in range(1, 11):
        ws.cell(row=r, column=col).border = THIN

r0 = 4 + len(CAT) + 1
ws.cell(row=r0,   column=2, value="REACHABILITY TEST").font = Font(bold=True, size=12)
ws.cell(row=r0+1, column=2, value="Test: something reachable inside 2 days, something aspirational at 2-4 weeks.")
head(ws, r0+2, ["", "Archetype", "Cheapest SKU", "Days", "Aspiration SKU", "Days", "Pass?"], None)
# (col of days-to-afford in the catalog rows, cheapest row, aspiration row)
tests = [("Grinder", "H", 7, 15), ("Dipper", "I", 4, 7), ("Whale", "J", 5, 13)]
for i, (name, dcol, cheap_r, asp_r) in enumerate(tests):
    r = r0 + 3 + i
    ws.cell(row=r, column=2, value=name).font = BOLD
    for col, f, fmt in [(3, f"=B{cheap_r}", "General"), (4, f"={dcol}{cheap_r}", "0.0"),
                        (5, f"=B{asp_r}", "General"),   (6, f"={dcol}{asp_r}", "0.0")]:
        c = ws.cell(row=r, column=col, value=f); c.fill, c.number_format = OUT, fmt
    c = ws.cell(row=r, column=7,
                value=f'=IF(AND(D{r}<=2,F{r}>=14,F{r}<=28),"PASS","FAIL")')
    c.fill, c.font = OUT, BOLD

# ── GUARDRAILS ─────────────────────────────────────────────────────────────────
ws = sheet("Guardrails")
ws["A1"] = "GUARDRAILS"; ws["A1"].font = Font(bold=True, size=13)
head(ws, 3, ["Guardrail", "Value", "Protects"], [34, 12, 62])
rows = [("Redemptions / player / week", 2, "Brand CAC budget + coupon inventory — the real control"),
        ("Daily coin earn ceiling", 900, "Farming. ~1.7x a Grinder's normal day: binds bots, not humans"),
        ("Voucher non-redemption rate", 0.15, "Assumed. Drives the coin-refund float below"),
        ("Whale IAP ARPDAU kill threshold", -0.02, "The thing this store was sold to protect")]
for r, (k, v, why) in enumerate(rows, start=4):
    ws.cell(row=r, column=1, value=k)
    c = ws.cell(row=r, column=2, value=v); c.fill = IN
    ws.cell(row=r, column=3, value=why)
ws["B6"].number_format = "0%"; ws["B7"].number_format = "0%"

ws["A9"] = "MINT VS SINK PER ARCHETYPE"; ws["A9"].font = Font(bold=True, size=12)
head(ws, 10, ["Archetype", "Mint/week", "Max sink/week", "Net", "Verdict"], [16, 13, 15, 12, 46])
for i, (name, earn_ref, avg_sku) in enumerate([("Grinder", "Earn!$F$10", "Catalog!$F$8"),
                                               ("Dipper",  "Earn!$F$11", "Catalog!$F$5"),
                                               ("Whale",   "Earn!$F$12", "Catalog!$F$6")]):
    r = 11 + i
    ws.cell(row=r, column=1, value=name).font = BOLD
    for col, f in [(2, f"={earn_ref}"), (3, f"={avg_sku}*$B$4"), (4, f"=B{r}-C{r}")]:
        c = ws.cell(row=r, column=col, value=f); c.fill, c.number_format = OUT, "#,##0"
    c = ws.cell(row=r, column=5, value=f'=IF(D{r}>0,"Mints faster than it sinks — needs a higher tier","Can sink faster than it mints — healthy")')
    c.fill = OUT

ws["A16"] = "Coins refunded/week from expiry"
ws["B16"] = "=SUM(C11:C13)*B6"; ws["B16"].fill, ws["B16"].number_format = OUT, "#,##0"

# ── UNIT ECONOMICS ─────────────────────────────────────────────────────────────
ws = sheet("UnitEconomics")
ws["A1"] = "UNIT ECONOMICS — the redemption rate is DERIVED from PlaySuper's published +5.2% ARPDAU, not guessed"
ws["A1"].font = Font(bold=True, size=13)
head(ws, 3, ["Line", "Value", "Source / formula"], [40, 16, 56])
UE = [
    ("DAU",                              1_000_000, "input", IN,  "#,##0"),
    ("Baseline ARPDAU",                       0.80, "input", IN,  '₹0.00'),
    ("Target ARPDAU lift",                   0.052, "PlaySuper published figure", IN, "0.0%"),
    ("Incremental revenue / DAU / day", "=B5*B6", "computed", OUT, '₹0.0000'),
    ("Average order value at redemption",      700, "input — I own this", IN, '₹#,##0'),
    ("PlaySuper commission %",                0.10, "input — I own this", IN, "0%"),
    ("PlaySuper commission ₹",         "=B8*B9", "computed", OUT, '₹#,##0'),
    ("Studio share of commission %",          0.50, "input — I own this", IN, "0%"),
    ("Studio revenue per redemption ₹","=B10*B11", "computed", OUT, '₹#,##0'),
    ("IMPLIED daily redemption rate",  "=B7/B12", "incremental ÷ studio rev per redemption", OUT, "0.000%"),
    ("Redemptions / day",              "=B4*B13", "computed", OUT, "#,##0"),
    ("Redemptions / month",            "=B14*30", "computed", OUT, "#,##0"),
    ("Average discount funded by brand ₹",     150, "input — I own this", IN, '₹#,##0'),
    ("BRAND CAC BUDGET REQUIRED / month","=B15*B16", "the real ceiling", OUT, '₹#,##0'),
    ("Share of DAU redeeming per month","=B15/B4", "computed", OUT, "0.0%"),
    ("Redemption cap headroom (cap 8/mo)","=8/(B18*30/30)", "cap ÷ modal monthly redemptions", OUT, "0.0"),
]
for i, (label, val, src, fill, fmt) in enumerate(UE):
    r = 4 + i
    ws.cell(row=r, column=1, value=label).font = BOLD if "IMPLIED" in label or "BRAND" in label else Font()
    c = ws.cell(row=r, column=2, value=val); c.fill, c.number_format = fill, fmt
    ws.cell(row=r, column=3, value=src)
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = THIN

ws["A23"] = "CONCLUSION 1"; ws["A23"].font = Font(bold=True, size=12)
ws["A24"] = ("Only ~0.12% of DAU — about 1 player in 840 — need to redeem daily to hit the published "
             "ARPDAU lift. So optimise for RETENTION BREADTH (many players holding a reserved Goal), "
             "not CONVERSION DEPTH (many checkouts). A held Goal costs the brand nothing and returns a "
             "player tomorrow; a checkout costs CAC and may not.")
ws["A26"] = "CONCLUSION 2"; ws["A26"].font = Font(bold=True, size=12)
ws["A27"] = ("The monthly brand CAC budget is the real ceiling — not studio cost. If brands fund less, "
             "the redemption cap tightens or the average discount drops. The coin faucet does not move. "
             "Faucet and budget are INDEPENDENT levers.")
for cell in ("A24", "A27"):
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[24].height = 58; ws.row_dimensions[27].height = 44

# ── VALUES (hardcoded, so Numbers/Preview never shows a blank model) ───────────
peg, win, loss, streak, wr = 10, 50, 15, 40, .5
arche = {"Grinder": (15, 7), "Dipper": (1.43, 2.5), "Whale": (6, 7)}  # matches/day, days played/wk
cpd = {k: m*(win*wr+loss*(1-wr))+streak*(d/7) for k, (m, d) in arche.items()}
ue = {}
ue["incr"] = 0.80*0.052
ue["studio_rev"] = 700*0.10*0.50
ue["rate"] = ue["incr"]/ue["studio_rev"]
ue["per_day"] = 1_000_000*ue["rate"]
ue["per_month"] = ue["per_day"]*30
ue["cac"] = ue["per_month"]*150

ws = sheet("Values")
ws["A1"] = "VALUES — hardcoded duplicate of every computed figure."
ws["A1"].font = Font(bold=True, size=13)
ws["A2"] = ("openpyxl writes formulas with no cached result, so Apple Numbers / Quick Look render those "
            "cells blank until the file is opened in Excel. This tab guarantees the model is readable "
            "anywhere. If it disagrees with the formula tabs, the formula tabs win.")
ws["A2"].alignment = Alignment(wrap_text=True); ws.row_dimensions[2].height = 42
ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16; ws.column_dimensions["D"].width = 16

vrows = [("PEG (coins per ₹1)", peg, "", ""),
         ("", "", "", ""),
         ("COINS/DAY", "Grinder", "Dipper", "Whale"),
         ("", round(cpd["Grinder"]), round(cpd["Dipper"]), round(cpd["Whale"])),
         ("48h budget", round(cpd["Grinder"]*2), round(cpd["Dipper"]*2), round(cpd["Whale"]*2)),
         ("3-week budget", round(cpd["Grinder"]*21), round(cpd["Dipper"]*21), round(cpd["Whale"]*21)),
         ("", "", "", ""),
         ("UNIT ECONOMICS", "", "", ""),
         ("Incremental rev / DAU / day", f"₹{ue['incr']:.4f}", "", ""),
         ("Studio rev per redemption", f"₹{ue['studio_rev']:.0f}", "", ""),
         ("Implied daily redemption rate", f"{ue['rate']*100:.3f}%", "", ""),
         ("Redemptions / day @ 1M DAU", f"{ue['per_day']:,.0f}", "", ""),
         ("Redemptions / month", f"{ue['per_month']:,.0f}", "", ""),
         ("Brand CAC budget / month", f"₹{ue['cac']:,.0f}", "", ""),
         ("Share of DAU redeeming / month", f"{ue['per_month']/1_000_000*100:.1f}%", "", "")]
for i, row in enumerate(vrows, start=4):
    for j, v in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=v)
        if j == 1 and isinstance(v, str) and v.isupper() and v:
            c.font = Font(bold=True)
        if i == 6:
            c.font = Font(bold=True)

ws["A21"] = "CATALOG (computed)"; ws["A21"].font = Font(bold=True)
head(ws, 22, ["SKU", "Coins", "Cash ₹", "Days: Grinder", "Days: Dipper", "Days: Whale"])
csv_rows = [["sku", "type", "mrp", "coin_cap", "coins", "cash", "days_grinder", "days_dipper", "days_whale"]]
for i, (n, name, typ, mrp, cap) in enumerate(CAT):
    coins = mrp*cap*peg
    cash  = mrp - coins/peg
    d = [coins/cpd[k] for k in ("Grinder", "Dipper", "Whale")]
    r = 23 + i
    for j, v in enumerate([name, round(coins), round(cash), round(d[0], 1), round(d[1], 1), round(d[2], 1)], start=1):
        ws.cell(row=r, column=j, value=v)
    csv_rows.append([name, typ, mrp, cap, round(coins), round(cash),
                     round(d[0], 1), round(d[1], 1), round(d[2], 1)])

wb.save(os.path.join(HERE, "economy_model.xlsx"))

with open(os.path.join(HERE, "economy_model.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerows(csv_rows)
    w.writerow([])
    w.writerow(["-- unit economics --"])
    for k, v in [("incremental_rev_per_dau_day", round(ue["incr"], 4)),
                 ("studio_rev_per_redemption", round(ue["studio_rev"])),
                 ("implied_daily_redemption_rate", f"{ue['rate']*100:.3f}%"),
                 ("redemptions_per_day_at_1m_dau", round(ue["per_day"])),
                 ("redemptions_per_month", round(ue["per_month"])),
                 ("brand_cac_budget_per_month", round(ue["cac"]))]:
        w.writerow([k, v])
    w.writerow([])
    w.writerow(["-- coins per day --"])
    for k, v in cpd.items():
        w.writerow([k, round(v)])

print("xlsx + csv written")
print(f"  implied redemption rate : {ue['rate']*100:.3f}% of DAU/day")
print(f"  redemptions/day @ 1M DAU: {ue['per_day']:,.0f}")
print(f"  brand CAC budget/month  : ₹{ue['cac']:,.0f}")
print(f"  coins/day               : " + ", ".join(f"{k} {v:.0f}" for k, v in cpd.items()))
